"""Spreadsheet — read and write CSV and Excel files."""


from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from harzoo.agent.kernel.tool import Tool, ToolResult, resolve_workspace_path, workspace_root_from

TOOL_VERSION = "2026-06-29"

MAX_ROWS = 500
MAX_COLS = 50


class SpreadsheetTool(Tool):
    """Read/write CSV and Excel (.xlsx) spreadsheets."""

    name = "Spreadsheet"
    description = (
        "Spreadsheet operations. Actions: read, write_csv, write_xlsx. "
        "Excel write requires: pip install openpyxl."
    )
    parameters = {
        "properties": {
            "action": {"type": "string", "enum": ["read", "write_csv", "write_xlsx"]},
            "file_path": {"type": "string", "description": "Path to CSV or XLSX file"},
            "rows": {
                "type": "array",
                "description": "Rows as array of arrays (write actions)",
                "items": {"type": "array", "items": {}},
            },
            "sheet": {"type": "string", "description": "Sheet name for xlsx", "default": "Sheet1"},
            "delimiter": {"type": "string", "description": "CSV delimiter", "default": ","},
        },
        "required": ["action", "file_path"],
    }

    def _read_csv(self, path: Path, delimiter: str) -> list[list[str]]:
        with path.open(encoding="utf-8-sig", newline="") as f:
            return [row for i, row in enumerate(csv.reader(f, delimiter=delimiter)) if i < MAX_ROWS]

    def _read_xlsx(self, path: Path, sheet: str) -> list[list[Any]]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:
            raise ImportError("openpyxl required: pip install openpyxl") from e
        wb = load_workbook(str(path), read_only=True, data_only=True)
        ws = wb[sheet] if sheet in wb.sheetnames else wb.active
        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                break
            rows.append([cell if cell is not None else "" for cell in row[:MAX_COLS]])
        wb.close()
        return rows

    def execute(
        self,
        action: str,
        file_path: str,
        rows: list[list[Any]] | None = None,
        sheet: str = "Sheet1",
        delimiter: str = ",",
        **kwargs: Any,
    ) -> ToolResult:
        act = str(action or "").strip().lower()
        root = workspace_root_from(kwargs.get("ctx"))
        if not str(file_path).strip():
            return ToolResult.failure("file_path must not be empty", code="INVALID_ARGUMENTS")
        path = resolve_workspace_path(str(file_path or "").strip(), root)

        if act == "read":
            if not path.is_file():
                return ToolResult.failure(f"File not found: {file_path}", code="PATH_NOT_FOUND")
            suffix = path.suffix.lower()
            try:
                if suffix == ".csv":
                    data = self._read_csv(path, str(delimiter or ",")[:1] or ",")
                elif suffix in (".xlsx", ".xlsm"):
                    data = self._read_xlsx(path, str(sheet or "Sheet1"))
                else:
                    return ToolResult.failure(f"Unsupported format: {suffix}", code="UNSUPPORTED_FORMAT")
            except ImportError as e:
                return ToolResult.failure(str(e), code="CAPABILITY_UNAVAILABLE")
            except Exception as e:
                return ToolResult.failure(f"{type(e).__name__}: {e}", code="READ_ERROR")
            return ToolResult.success({"file_path": str(path), "row_count": len(data), "rows": data})

        if act == "write_csv":
            if not rows:
                return ToolResult.failure("rows required for write_csv", code="INVALID_ARGUMENTS")
            path.parent.mkdir(parents=True, exist_ok=True)
            with path.open("w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f, delimiter=str(delimiter or ",")[:1] or ",")
                for row in rows[:MAX_ROWS]:
                    writer.writerow([str(c) for c in row[:MAX_COLS]])
            return ToolResult.success({"file_path": str(path), "rows_written": min(len(rows), MAX_ROWS), "saved": True})

        if act == "write_xlsx":
            if not rows:
                return ToolResult.failure("rows required for write_xlsx", code="INVALID_ARGUMENTS")
            try:
                from openpyxl import Workbook  # type: ignore
            except ImportError:
                return ToolResult.failure("openpyxl required: pip install openpyxl", code="CAPABILITY_UNAVAILABLE")
            wb = Workbook()
            ws = wb.active
            ws.title = str(sheet or "Sheet1")[:31]
            for row in rows[:MAX_ROWS]:
                ws.append([c for c in row[:MAX_COLS]])
            path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(str(path))
            return ToolResult.success({"file_path": str(path), "rows_written": min(len(rows), MAX_ROWS), "saved": True})

        return ToolResult.failure(f"Unknown action: {action}", code="INVALID_ARGUMENTS")


TOOL = SpreadsheetTool
